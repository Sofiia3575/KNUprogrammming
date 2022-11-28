import os
import sys
from openpyxl import *
import re



def output_cells(folder, text):

    #os.chdir(folder)
    rgx = re.compile(text)

    for path, dirs, files in os.walk(folder):


        for item in files:
            if not item.endswith(".xlsx"):
                continue
            fullname = os.path.join(path, item)
            wb = load_workbook(fullname)
            try:
                wb = load_workbook(fullname)
                lst = wb.get_sheet_names()
                for ws in wb:
                    #ws = wb[sheet]
                    for i, row in enumerate(ws.iter_rows()):

                        for j, cell in enumerate(row):
                            cell_text = cell.value
                            if not cell_text:
                                continue
                            if re.match(cell_text, text):
                                title = ws.title

                                print(f"{fullname}, title{i}{j} has text {text} ")

            except:
                print(f"can\'t open {path}")





if __name__ == '__main__':
    if len(sys.argv) >= 3:
        folder = sys.argv[1]
        text = sys.argv[2]

    else:
        folder = input("folder:")
        text = input("to find:")

    output_cells(folder, text)
